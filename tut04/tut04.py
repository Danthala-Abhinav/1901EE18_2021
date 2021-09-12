import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
def output_by_subject():
    try:
        os.makedirs('output_by_subject')      
    except:
        pass
    with open('regtable_old.csv', 'r') as file:
        reader = csv.reader(file, delimiter=',')
        for row in reader:
            del row[4:8]
            del row[2]
            if(row[2]=='subno'):
                continue
            else:
                try:
                    with open(f'output_by_subject\\{row[2]}.xlsx'):
                        wb = load_workbook(r'output_by_subject\\{}.xlsx'.format(row[2]))
                        sheet = wb.active
                        line = [row[0],row[1],row[2],row[3]]
                        sheet.append(line)
                        wb.save(r'output_by_subject\\{}.xlsx'.format(row[2]))
                except:    
                        wb = Workbook()
                        sheet = wb.active
                        line1 = ['rollno','register_sem','subno','sub_type']
                        line2 = [row[0],row[1],row[2],row[3]]
                        sheet.append(line1)
                        sheet.append(line2)
                        wb.save(f'output_by_subject\\{row[2]}.xlsx')
    return

def output_individual_roll():
    try:
        os.makedirs('output_individual_roll')       
    except:
        pass
    with open('regtable_old.csv', 'r') as file:
        reader = csv.reader(file, delimiter=',')
        for row in reader:
            del row[4:8]
            del row[2]
            if(row[0]=='rollno'):
                continue
            else:
                try:
                    with open(f'output_individual_roll\\{row[0]}.xlsx'):
                        wb = load_workbook(r'output_individual_roll\\{}.xlsx'.format(row[0]))
                        sheet = wb.active
                        line = [row[0],row[1],row[2],row[3]]
                        sheet.append(line)
                        wb.save(r'output_individual_roll\\{}.xlsx'.format(row[0]))
                except:    
                        wb = Workbook()
                        sheet = wb.active
                        line1 = ['rollno','register_sem','subno','sub_type']
                        line2 = [row[0],row[1],row[2],row[3]]
                        sheet.append(line1)
                        sheet.append(line2)
                        wb.save(f'output_individual_roll\\{row[0]}.xlsx')
    return

output_by_subject()
output_individual_roll()