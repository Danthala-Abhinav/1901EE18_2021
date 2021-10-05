import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook

#Grades mapping
grade_dict = {'AA':10,'AB':9,'BB':8,'BC':7,'CC':6,'CD':5,'DD':4,'F':0,'I':0,
		      'AA*':10,'AB*':9,'BB*':8,'BC*':7,'CC*':6,'CD*':5,'DD*':4,'F*':0,'I*':0}

def Result_generator(Spi,sheet,Credits):
#Obtaining final results
	sheet.append(["Semester No"]+[x for x in range(1,9)])
	sheet.append(["Semester wise Credit taken"] + Credits)
	sheet.append(["SPI"] + Spi)
	creditsprefix = [sum(Credits[:i+1]) for i in range(len(Credits))]
	sheet.append(["Total Credits taken"]+creditsprefix)
	CPI_num=[Spi[i]*Credits[i] for i in range(len(Credits))] 
	sheet.append(["CPI"]+[round(sum(CPI_num[:i+1])/creditsprefix[i],2) for i in range(len(Spi))])
	return
	

def Spi_Credits(wb,grade_dict):
#Obtaining Spi and Credits 
	Credits= []
	Spi=[]
	for sheet in wb.sheetnames[1:]:
		ws = wb[sheet]
		credits = [int(cell.value) for cell in ws["E"][1:]]
		spi = [grade_dict[cell.value.strip().strip("*")] for cell in ws["G"][1:]]
		Credits.append(sum(credits))
		Spi.append(round ( sum([credits[i]*spi[i] for i in range(len(spi))]) / sum(credits),2 ))
	return Spi,Credits

#Generating 'Overall' sheet
def sheet_for_overall():
	with open('names-roll.csv','r') as file:
		reader = csv.reader(file)
		namelist = [list(row) for row in reader][1:]
	for row in namelist:
		name = row[1]
		roll = row[0]
		wb = load_workbook(r'output\\{}.xlsx'.format(roll))
		sheet = wb.active
		sheet.column_dimensions["A"].width = 30
		sheet.append(["RollNo",roll])
		sheet.append(["Name of Student",name])
		sheet.append(["Discipline",roll[4:6]])

		#Obtaining Spi and Credits from Spi_Credits() function
		Spi,Credits= Spi_Credits(wb,grade_dict)

		#Obtaining final results
		Result_generator(Spi,sheet,Credits)
		wb.save(r'output\\{}.xlsx'.format(roll))


#Creating a dictionary of subjects with key as subject code and its data as remaining in row
def subjects_dict():
	with open('subjects_master.csv','r') as file:
		reader = csv.reader(file)
		subjects_csv = [list(row) for row in reader][1:]
	sub_dict = {}
	maxcol_width = 0
	for row in subjects_csv:
		sub = row[0]
		if sub not in sub_dict:
			sub_dict[sub] = row[1:]
			maxcol_width = max(maxcol_width , len(row[1]))
	return sub_dict , maxcol_width


#Check for existence of sheet
def sheetexist(sheetname , wb):
	for sheet in wb.sheetnames:
		if (sheetname==sheet):
			return 1
	return 0


def generate_marksheet():
	try:
		os.makedirs('output')
		#Creating a directory
	except:
		pass
	sub_dict = []
	#From subjects_dict() function
	sub_dict , maxcolumn_width = subjects_dict()
	with open('grades.csv', 'r') as file:
		grades_csv = csv.reader(file)
		grade_list = [list(row) for row in grades_csv][1:]
	for row in grade_list:
		roll = row[0]
		sem = row[1]
		semno = int(sem)
		filepath = './output/'+'{}.xlsx'.format(roll)
		if not os.path.isfile(filepath):
			#Creating 'Overall' sheet
			wb = Workbook()
			sheet1 = wb.active
			sheet1.title = 'Overall'
			wb.save(r'output\\{}.xlsx'.format(row[0]))
		wb = load_workbook(r'output\\{}.xlsx'.format(roll))

		#Check for existence of sheet
		if not sheetexist(f'Sem{sem}',wb):
			#Creating respective sem sheets
			wb.create_sheet(f'Sem{sem}',semno)
			sheet = wb[f'Sem{semno}']
			sheet.column_dimensions['C'].width = maxcolumn_width
		sheet = wb[f'Sem{semno}']

		if (sheet.max_row==1):
			#If nothing is there in sem sheet,then append headings
			sheet.append(['Sl No.','Subject No.','Subject Name','L-T-P','Credit','Subject Type','Grade']) 
		SubCode = row[2]
		Credit = row[3]
		Grade = row[4]
		Sub_Type = row[5]
		SubName,LTP,crd = sub_dict[f"{SubCode}"][0],sub_dict[f"{SubCode}"][1],sub_dict[f"{SubCode}"][2]
		#Appending data to sem sheets
		sheet.append([sheet.max_row,SubCode,SubName,LTP,crd,Sub_Type,Grade])
		wb.save(r'output\\{}.xlsx'.format(roll))      
	#Generating 'Overall' sheet
	sheet_for_overall()
	return

generate_marksheet()